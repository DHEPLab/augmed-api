import os
from io import BytesIO
from typing import List

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from src.common.exception.BusinessException import (BusinessException,
                                                    BusinessExceptionEnum)
from src.user.model.configuration import Configuration


class ExcelConfigurationParser:
    def __init__(self, excel_stream: BytesIO):
        self.workbook = load_workbook(filename=excel_stream)
        self.sheet = self.workbook.active

    def parse(self) -> List[Configuration]:
        configurations = []
        headers = [cell.value for cell in self.sheet[1]]
        user_index = headers.index("User")
        case_index = headers.index("Case No.")

        current_config = None

        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row):
            if not any(cell.value is not None for cell in row):
                continue
            row_data = [cell.value for cell in row]

            user, case = self._validate_and_extract_user_case(
                row, user_index, case_index
            )
            case_id = self._validate_and_convert_case_id(case)

            if self._should_create_new_config(
                current_config, user, case_id, row, user_index, case_index
            ):
                current_config = self._create_new_config(user, case_id)
                configurations.append(current_config)

            self._process_path_config(current_config, row_data, headers)

        return configurations

    def _validate_and_extract_user_case(self, row, user_index, case_index):
        user_cell = row[user_index]
        case_cell = row[case_index]

        user = self._resolve_merged_cells(user_cell)
        if not user:
            raise BusinessException(BusinessExceptionEnum.InvalidUserEmail)

        case = self._resolve_merged_cells(case_cell)
        if not case:
            raise BusinessException(BusinessExceptionEnum.InvalidCaseId)

        return user, case

    def _validate_and_convert_case_id(self, case):
        try:
            return int(case)
        except (ValueError, TypeError):
            raise BusinessException(BusinessExceptionEnum.InvalidCaseId)

    def _should_create_new_config(
        self, current_config, user, case_id, row, user_index, case_index
    ):
        is_merged_user = self._is_merged_cell(row[user_index])
        is_merged_case = self._is_merged_cell(row[case_index])

        return (
            not current_config
            or current_config.user_email != user
            or current_config.case_id != case_id
            or not (is_merged_user and is_merged_case)
        )

    def _create_new_config(self, user, case_id):
        return Configuration(user_email=user, case_id=case_id, path_config=[])

    def _process_path_config(self, current_config, row_data, headers):
        path = row_data[headers.index("Path")]
        collapse = row_data[headers.index("Collapse")]
        highlight = row_data[headers.index("Highlight")]
        top = row_data[headers.index("Top")]

        if path:
            style = self._build_style_dict(collapse, highlight, top)
            if style:
                current_config.path_config.append({"path": path, "style": style})

    def _resolve_merged_cells(self, cell):
        if cell.value is not None:
            return cell.value
        for range_ in self.sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                min_col, min_row, _, _ = range_boundaries(str(range_))
                return self.sheet.cell(row=min_row, column=min_col).value
        return None

    def _is_merged_cell(self, cell):
        for range_ in self.sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                return True
        return False

    def _build_style_dict(self, collapse, highlight, top) -> dict:
        style = {}
        if collapse is not None:
            style["collapse"] = collapse
        if highlight is not None:
            style["highlight"] = highlight
        if top is not None:
            style["top"] = top
        return style


def parse_excel_stream_to_configurations(excel_stream: BytesIO) -> List[Configuration]:
    parser = ExcelConfigurationParser(excel_stream)
    return parser.parse()


def is_excel_file(filename: str | None):
    if filename is None:
        return False

    _, extension = os.path.splitext(filename)
    return extension in [".xlsx", ".xls"]
