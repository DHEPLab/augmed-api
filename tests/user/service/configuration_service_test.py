from io import BytesIO

import pytest
from openpyxl import Workbook
from werkzeug.exceptions import InternalServerError

from src.common.exception.BusinessException import BusinessException, BusinessExceptionEnum
from src.user.repository.configuration_repository import ConfigurationRepository
from src.user.service.configuration_service import ConfigurationService


@pytest.fixture
def mock_repo(mocker):
    return mocker.Mock(ConfigurationRepository)


@pytest.fixture
def valid_excel_file():
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'User'
    ws['B1'] = 'Case No.'
    ws['C1'] = 'Path'
    ws['D1'] = 'Collapse'
    ws['E1'] = 'Highlight'
    # Add some dummy data
    ws.append(['usera@example.com', '1', 'Background.abc', 'TRUE', 'TRUE'])
    ws.append(['userb@example.com', '2', 'Background.xyz', 'FALSE', 'TRUE'])
    # Save the workbook to a BytesIO object
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)  # Rewind the stream to the beginning
    return excel_stream


@pytest.fixture
def config_data():
    return [
        {"user_email": "usera@example.com", "case_id": 1, "path_config": [
            {"path": "Background.abc", "style": {"Collapse": True, "Highlight": True}}
        ]},
        {"user_email": "userb@example.com", "case_id": 2, "path_config": [
            {"path": "Background.xyz", "style": {"Collapse": False, "Highlight": True}}
        ]}
    ]


def test_process_excel_file_success(mocker, mock_repo, valid_excel_file, config_data):
    # Setup the mock to return our prepared config data
    mocker.patch('src.user.utils.excel_parser.parse_excel_stream_to_configurations', return_value=config_data)
    service = ConfigurationService(repository=mock_repo)

    response = service.process_excel_file(valid_excel_file)
    print('response',response.__str__())
    # Assertions to check if the response is as expected
    assert len(response) == len(config_data)
    assert response[0]["user_case_key"] == "usera@example.com-1"
    assert response[0]["status"] == "added"
    assert response[1]["user_case_key"] == "userb@example.com-2"
    assert response[1]["status"] == "added"
    mock_repo.clean_configurations.assert_called_once()
    assert mock_repo.save_configuration.call_count == len(config_data)



def test_parser_error(mocker, mock_repo, valid_excel_file):
    mocker.patch('src.user.service.configuration_service.parse_excel_stream_to_configurations',
                 side_effect=Exception("Parse failed"))
    service = ConfigurationService(repository=mock_repo)

    with pytest.raises(BusinessException) as exc_info:
        service.process_excel_file(valid_excel_file)

    assert exc_info.value.error == BusinessExceptionEnum.ConfigFileIncorrect


def test_database_cleaning_error(mocker, mock_repo, valid_excel_file):
    mocker.patch('src.user.utils.excel_parser.parse_excel_stream_to_configurations', return_value=[])
    mock_repo.clean_configurations.side_effect = Exception("Database cleanup failed")
    service = ConfigurationService(repository=mock_repo)

    with pytest.raises(InternalServerError):
        service.process_excel_file(valid_excel_file)


def test_database_save_error(mocker, mock_repo, valid_excel_file):
    config_data = [{"user_email": "usera@example.com", "case_id": 1, "path_config": []}]
    mocker.patch('src.user.utils.excel_parser.parse_excel_stream_to_configurations', return_value=config_data)
    mock_repo.clean_configurations.return_value = None
    mock_repo.save_configuration.side_effect = Exception("Save failed")
    service = ConfigurationService(repository=mock_repo)

    response = service.process_excel_file(valid_excel_file)

    assert response[0]["user_case_key"] == 'usera@example.com-1'
    assert response[0]["status"] == 'failed'
    mock_repo.clean_configurations.assert_called_once()
