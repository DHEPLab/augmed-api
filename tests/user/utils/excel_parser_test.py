from io import BytesIO

import pytest
from openpyxl import Workbook

from src.common.exception.BusinessException import BusinessException
from src.user.utils.excel_parser import is_excel_file, parse_excel_stream_to_configurations


def test_should_parse_excell_stream_correctly_when_all_config_are_set():
    # Prepare the test data (Workbook with multiple configurations)
    wb = Workbook()
    ws = wb.active
    # Headers
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    # Data for multiple users and cases
    ws.append(['usera@example.com', '1', 'Background.abc', True, True, 1])
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)

    ws.append([None, None, 'background.xxx', True, False, 2])
    ws.append(['userb@example.com', '1', 'Background.patient demo', False, False, 1.5])
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    result = parse_excel_stream_to_configurations(excel_stream)

    #result_dicts = [config.__dict__ for config in result]
    result_dicts = [config.to_dict()for config in result]

    # Define expected results
    expected_results = [
        {
            'user_email': 'usera@example.com',
            'case_id': 1,
            'path_config': [
                {'path': 'Background.abc', 'style': {'collapse': True, 'highlight': True, 'top': 1}},
                {'path': 'background.xxx', 'style': {'collapse': True, 'highlight': False, 'top': 2}}
            ]
        },
        {
            'user_email': 'userb@example.com',
            'case_id': 1,
            'path_config': [
                {'path': 'Background.patient demo', 'style': {'collapse': False, 'highlight': False, "top": 1.5}}
            ]
        }
    ]

    # Assert with a simple comparison
    assert result_dicts == expected_results


def test_should_ignore_none_config():
    # Prepare the test data (Workbook with multiple configurations)
    wb = Workbook()
    ws = wb.active
    # Headers
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    # Data for multiple users and cases
    ws.append(['usera@example.com', '1', 'Background.abc', None, True, None])
    ws.append([None, None, 'background.xxx', True, None, None])
    ws.append([None, None, 'Background.patient demo', None, None, None])
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    result = parse_excel_stream_to_configurations(excel_stream)

    # Check if the result matches the expected configuration
    assert len(result) == 1
    assert result[0].user_email == 'usera@example.com'
    assert result[0].case_id == 1
    assert len(result[0].path_config) == 2

    # Assert the configuration of paths
    expected_path_0 = {
        'path': 'Background.abc',
        'style': {'highlight': True}
    }
    expected_path_1 = {
        'path': 'background.xxx',
        'style': {'collapse': True}
    }

    # Check each path configuration for correctness
    assert result[0].path_config[0] == expected_path_0
    assert result[0].path_config[1] == expected_path_1


def test_should_ignore_none_config_while_keep_user_case_relationship():
    # Prepare the test data (Workbook with multiple configurations)
    wb = Workbook()
    ws = wb.active
    # Headers
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    # Data for multiple users and cases
    ws.append(['usera@example.com', '1', 'Background.patient demo', None, None, None])
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    result = parse_excel_stream_to_configurations(excel_stream)
    assert len(result) == 1
    assert result[0].user_email == 'usera@example.com'
    assert result[0].case_id == 1
    assert len(result[0].path_config) == 0


def test_should_ignore_blank_line():
    # Prepare the test data (Workbook with multiple configurations)
    wb = Workbook()
    ws = wb.active
    # Headers
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    # Data for multiple users and cases
    ws.append(['usera@example.com', '1', 'Background.patient demo', None, None])
    ws.append([None, None, None, None, None])
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    result = parse_excel_stream_to_configurations(excel_stream)

    assert len(result) == 1
    assert result[0].user_email == 'usera@example.com'
    assert result[0].case_id == 1
    assert len(result[0].path_config) == 0


def test_should_keep_duplicate_user_case_relationship():
    # Prepare the test data (Workbook with multiple configurations)
    wb = Workbook()
    ws = wb.active
    # Headers
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    # Data for multiple users and cases
    ws.append(['usera@example.com', '1', 'Background.patient demo', None, None, None])
    ws.append([None, None, 'Background.drug', None, None, None])
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.append(['usera@example.com', '1', 'Background.patient demo', None, None, None])

    ws.append(['usera@example.com', '1', None, None, None, None])

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    result = parse_excel_stream_to_configurations(excel_stream)
    assert len(result) == 3
    assert result[0].user_email == 'usera@example.com'
    assert result[0].case_id == 1
    assert len(result[0].path_config) == 0

    assert result[1].user_email == 'usera@example.com'

    assert result[1].case_id == 1
    assert len(result[1].path_config) == 0
    assert result[1].user_email == 'usera@example.com'

    assert result[2].case_id == 1
    assert len(result[1].path_config) == 0
    assert result[2].user_email == 'usera@example.com'


def test_invalid_user_email_raises_exception():
    wb = Workbook()
    ws = wb.active
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    ws.append([None, '1', 'Background.patient demo', None, None, None])  # Invalid user

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    with pytest.raises(BusinessException) as excinfo:
        parse_excel_stream_to_configurations(excel_stream)

    assert "Invalid user email in config file." in str(excinfo.value.error.value)


def test_invalid_case_id_raises_exception():
    wb = Workbook()
    ws = wb.active
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    ws.append(['usera@example.com', 'abc', 'Background.patient demo', None, None, None])  # Invalid case ID

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    with pytest.raises(BusinessException) as excinfo:
        parse_excel_stream_to_configurations(excel_stream)

    assert "Invalid case id in config file." in str(excinfo.value.error.value)


def test_is_excel_file():
    excels = ['test.xlsx', 'test.xls']
    not_excels = ['test.numbers', 'test.txt', '', None]

    for x in excels:
        assert is_excel_file(x) is True
    for x in not_excels:
        assert is_excel_file(x) is False


def test_invalid_non_number_top_config_should_raises_exception():
    wb = Workbook()
    ws = wb.active
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    ws.append(['uset@test.com', '1', 'Background.patient demo', None, None, 'NAN'])

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    with pytest.raises(BusinessException) as excinfo:
        parse_excel_stream_to_configurations(excel_stream)

    assert "Error while processing Excel file, please check again." in str(excinfo.value.error.value)


def test_invalid_top_config_on_root_node_should_raises_exception():
    wb = Workbook()
    ws = wb.active
    ws.append(['User', 'Case No.', 'Path', 'Collapse', 'Highlight', 'Top'])
    ws.append(['uset@test.com', '1', 'Background', None, None, 1])

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    with pytest.raises(BusinessException) as excinfo:
        parse_excel_stream_to_configurations(excel_stream)

    assert "Error while processing Excel file, please check again." in str(excinfo.value.error.value)
